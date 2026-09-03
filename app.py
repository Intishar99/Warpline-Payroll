import os
import io
import zipfile
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import check_password_hash
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from models import db, PayPeriod, TimeEntry, CampaignEntry, AdminUser, Campaign, CampaignTier
from parsers import parse_raw_payroll_csv
from calc_engine import calc_agent_payroll
from paystub_generator import generate_paystub_docx

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-key-change-in-production")

database_url = os.environ.get("DATABASE_URL")
if database_url:
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
else:
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{os.path.join(BASE_DIR, 'payroll.db')}"

app.config["UPLOAD_FOLDER"] = os.path.join(BASE_DIR, "uploads")

db.init_app(app)
csrf = CSRFProtect(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "Please log in to access WarpLine Payroll."
login_manager.login_message_category = "error"


@login_manager.user_loader
def load_user(user_id):
    return AdminUser.query.get(int(user_id))


# Pre-filled from the WarpLine Agent Compensation Structure PDF. These are seeded once,
# on first run, and are fully editable afterward via the Manage Campaigns page -- nothing
# here is hardcoded into the calculation logic itself, this is just the starting data.
PDF_CAMPAIGNS = {
    "Solar": {
        "sit_commission": 15, "sale_commission": 35,
        "tiers": [(0,3.00),(6,4.00),(8,5.00),(16,6.00),(24,6.50),(30,7.00),(38,7.50)],
    },
    "Roofing IL/IA (Insurance)": {
        "sit_commission": 10, "sale_commission": 10,
        "tiers": [(0,3.00),(8,4.00),(12,5.00),(22,6.00),(28,6.50),(36,7.00),(48,7.50)],
    },
    "Roofing OK/TX (Insurance)": {
        "sit_commission": 10, "sale_commission": 10,
        "tiers": [(0,3.00),(7,4.00),(11,5.00),(21,6.00),(27,6.50),(35,7.00),(47,7.50)],
    },
    "Roofing (Retail)": {
        "sit_commission": 15, "sale_commission": 25,
        "tiers": [(0,3.00),(6,4.00),(8,5.00),(12,5.50),(16,6.00),(24,6.50),(30,7.00),(38,7.50)],
    },
}


def seed_campaigns():
    """Runs once -- if no campaigns exist yet, pre-fill the four from the compensation
    PDF. Idempotent: does nothing if campaigns already exist (e.g. after an admin has
    already started editing them)."""
    if Campaign.query.first() is not None:
        return
    for name, data in PDF_CAMPAIGNS.items():
        campaign = Campaign(name=name, sit_commission=data["sit_commission"], sale_commission=data["sale_commission"])
        db.session.add(campaign)
        db.session.flush()
        for threshold, rate in data["tiers"]:
            db.session.add(CampaignTier(campaign_id=campaign.id, appointments_threshold=threshold, hourly_rate=rate))
    db.session.commit()


with app.app_context():
    db.create_all()
    seed_campaigns()


def make_label(start_date, end_date):
    if start_date.month == end_date.month:
        return f"{start_date.strftime('%b %-d')} - {end_date.strftime('%-d, %Y')}"
    return f"{start_date.strftime('%b %-d')} - {end_date.strftime('%b %-d, %Y')}"


def get_agent_results(period):
    time_entries = TimeEntry.query.filter_by(pay_period_id=period.id).order_by(TimeEntry.agent_name).all()
    campaign_entries = CampaignEntry.query.filter_by(pay_period_id=period.id).all()

    campaigns_by_agent = {}
    for ce in campaign_entries:
        campaigns_by_agent.setdefault(ce.agent_name, []).append({
            "campaign": ce.campaign,
            "appointments": ce.appointments,
            "valid_sits": ce.valid_sits,
            "sales": ce.sales,
        })

    payroll_rows = []
    grand_total = 0.0
    for te in time_entries:
        entries = campaigns_by_agent.get(te.agent_name, [])
        result = calc_agent_payroll(
            te.total_hours, entries,
            hourly_rate=te.hourly_rate, manual_hours=te.manual_hours, spiffs=te.spiffs,
        )
        payroll_rows.append({"name": te.agent_name, "total_hours": te.total_hours, **result})
        grand_total += result["gross_pay"]

    return payroll_rows, round(grand_total, 2)


# ---------- Auth ----------

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = AdminUser.query.filter_by(username=username).first()
        if user is None or not check_password_hash(user.password_hash, password):
            flash("Incorrect username or password.", "error")
            return redirect(url_for("login"))

        login_user(user)
        next_page = request.args.get("next")
        return redirect(next_page or url_for("index"))

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ---------- Campaign management ----------

@app.route("/campaigns")
@login_required
def campaigns_page():
    campaigns = Campaign.query.order_by(Campaign.name).all()
    return render_template("campaigns.html", campaigns=campaigns)


@app.route("/campaigns/new", methods=["POST"])
@login_required
def new_campaign():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Campaign name is required.", "error")
        return redirect(url_for("campaigns_page"))
    if Campaign.query.filter_by(name=name).first():
        flash(f"A campaign named '{name}' already exists.", "error")
        return redirect(url_for("campaigns_page"))

    campaign = Campaign(name=name, sit_commission=0.0, sale_commission=0.0)
    db.session.add(campaign)
    db.session.commit()
    flash(f"Campaign '{name}' created. Add its tier rows and commission rates below.", "success")
    return redirect(url_for("edit_campaign", campaign_id=campaign.id))


@app.route("/campaigns/<int:campaign_id>/edit", methods=["GET", "POST"])
@login_required
def edit_campaign(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)

    if request.method == "POST":
        campaign.name = request.form.get("name", campaign.name).strip()
        campaign.sit_commission = float(request.form.get("sit_commission") or 0)
        campaign.sale_commission = float(request.form.get("sale_commission") or 0)

        CampaignTier.query.filter_by(campaign_id=campaign.id).delete()
        thresholds = request.form.getlist("tier_threshold")
        rates = request.form.getlist("tier_rate")
        for t, r in zip(thresholds, rates):
            if t.strip() and r.strip():
                db.session.add(CampaignTier(
                    campaign_id=campaign.id,
                    appointments_threshold=int(t),
                    hourly_rate=float(r),
                ))
        db.session.commit()
        flash(f"'{campaign.name}' saved.", "success")
        return redirect(url_for("campaigns_page"))

    tiers = CampaignTier.query.filter_by(campaign_id=campaign.id).order_by(CampaignTier.appointments_threshold).all()
    return render_template("edit_campaign.html", campaign=campaign, tiers=tiers)


@app.route("/campaigns/<int:campaign_id>/delete", methods=["POST"])
@login_required
def delete_campaign(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    if CampaignEntry.query.filter_by(campaign_id=campaign.id).first():
        flash(f"Can't delete '{campaign.name}' -- it has payroll data attached to it.", "error")
        return redirect(url_for("campaigns_page"))
    db.session.delete(campaign)
    db.session.commit()
    flash(f"Campaign deleted.", "success")
    return redirect(url_for("campaigns_page"))


# ---------- Pay periods ----------

@app.route("/")
@login_required
def index():
    periods = PayPeriod.query.order_by(PayPeriod.created_at.desc()).all()
    return render_template("index.html", periods=periods)


@app.route("/period/new", methods=["POST"])
@login_required
def new_period():
    start_str = request.form.get("start_date", "").strip()
    end_str = request.form.get("end_date", "").strip()

    if not start_str or not end_str:
        flash("Please select both a start and end date for this pay period.", "error")
        return redirect(url_for("index"))

    try:
        start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
    except ValueError:
        flash("Couldn't read those dates -- please try again.", "error")
        return redirect(url_for("index"))

    if end_date < start_date:
        flash("End date can't be before the start date.", "error")
        return redirect(url_for("index"))

    period = PayPeriod(label=make_label(start_date, end_date), start_date=start_date, end_date=end_date)
    db.session.add(period)
    db.session.commit()
    return redirect(url_for("upload_page", period_id=period.id))


@app.route("/period/<int:period_id>/upload", methods=["GET", "POST"])
@login_required
def upload_page(period_id):
    period = PayPeriod.query.get_or_404(period_id)

    if request.method == "POST":
        file = request.files.get("raw_csv")
        if not file or file.filename == "":
            flash("Please choose a raw dialer CSV to upload.", "error")
            return redirect(url_for("upload_page", period_id=period_id))

        try:
            df = parse_raw_payroll_csv(file)
        except Exception as e:
            flash(f"Couldn't read that file: {e}", "error")
            return redirect(url_for("upload_page", period_id=period_id))

        TimeEntry.query.filter_by(pay_period_id=period.id).delete()
        for _, row in df.iterrows():
            db.session.add(TimeEntry(
                pay_period_id=period.id,
                agent_name=row["Name"],
                break_hours=row["Break (t)"], training_hours=row["Training (t)"],
                lunch_hours=row["Lunch (t)"], manual_dial_hours=row["Manual Dial (t)"],
                talk_hours=row["Ready:Talk Time"], wait_hours=row["Ready:Wait Time"],
                wrap_hours=row["Ready:Wrap Time"], total_hours=row["Total Hours"],
            ))
        db.session.commit()
        flash(f"Loaded {len(df)} agents from the raw file.", "success")
        return redirect(url_for("entries_page", period_id=period.id))

    return render_template("upload.html", period=period)


@app.route("/period/<int:period_id>/entries", methods=["GET", "POST"])
@login_required
def entries_page(period_id):
    period = PayPeriod.query.get_or_404(period_id)
    time_entries = TimeEntry.query.filter_by(pay_period_id=period.id).order_by(TimeEntry.agent_name).all()
    campaigns = Campaign.query.order_by(Campaign.name).all()

    if not time_entries:
        flash("Upload a raw dialer CSV first.", "error")
        return redirect(url_for("upload_page", period_id=period_id))
    if not campaigns:
        flash("No campaigns set up yet -- add one on the Manage Campaigns page first.", "error")
        return redirect(url_for("campaigns_page"))

    if request.method == "POST":
        CampaignEntry.query.filter_by(pay_period_id=period.id).delete()

        for te in time_entries:
            safe_name = te.agent_name.replace(" ", "_")

            rate_val = request.form.get(f"{safe_name}__hourly_rate", "").strip()
            spiffs_val = request.form.get(f"{safe_name}__spiffs", "").strip()
            manual_hours_val = request.form.get(f"{safe_name}__manual_hours", "").strip()
            te.hourly_rate = float(rate_val) if rate_val else None
            te.spiffs = float(spiffs_val) if spiffs_val else 0.0
            te.manual_hours = float(manual_hours_val) if manual_hours_val else 0.0

            for campaign in campaigns:
                appt = request.form.get(f"{safe_name}__{campaign.id}__appointments", "").strip()
                sits = request.form.get(f"{safe_name}__{campaign.id}__sits", "").strip()
                sales = request.form.get(f"{safe_name}__{campaign.id}__sales", "").strip()

                if appt or sits or sales:
                    db.session.add(CampaignEntry(
                        pay_period_id=period.id,
                        agent_name=te.agent_name,
                        campaign_id=campaign.id,
                        appointments=int(appt) if appt else 0,
                        valid_sits=int(sits) if sits else 0,
                        sales=int(sales) if sales else 0,
                    ))
        db.session.commit()
        flash("Campaign data saved.", "success")
        return redirect(url_for("results_page", period_id=period.id))

    existing = {}
    for ce in CampaignEntry.query.filter_by(pay_period_id=period.id).all():
        existing.setdefault(ce.agent_name, {})[ce.campaign_id] = ce

    return render_template(
        "entries.html", period=period, time_entries=time_entries,
        campaigns=campaigns, existing=existing,
    )


@app.route("/period/<int:period_id>/results")
@login_required
def results_page(period_id):
    period = PayPeriod.query.get_or_404(period_id)
    payroll_rows, grand_total = get_agent_results(period)
    return render_template("results.html", period=period, payroll_rows=payroll_rows, grand_total=grand_total)


@app.route("/period/<int:period_id>/export.xlsx")
@login_required
def export_xlsx(period_id):
    period = PayPeriod.query.get_or_404(period_id)
    time_entries = {te.agent_name: te for te in TimeEntry.query.filter_by(pay_period_id=period.id).all()}
    payroll_rows, grand_total = get_agent_results(period)
    campaigns = Campaign.query.order_by(Campaign.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="0E3A52", end_color="0E3A52", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    money_fmt = "$#,##0.00"

    headers = (
        ["Name", "Break (t)", "Training (t)", "Lunch (t)", "Manual Dial (t)",
         "Ready:Talk Time", "Ready:Wait Time", "Ready:Wrap Time", "Manual Hours", "Total Hours",
         "Hourly Rate (USD)", "Total Hours Pay"]
        + [f"{c.name} Appts" for c in campaigns]
        + [f"{c.name} Sits" for c in campaigns]
        + [f"{c.name} Sales" for c in campaigns]
        + ["Total Appointments", "Commission Total", "Spiffs", "Amount in USD"]
    )
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_idx = 2
    for row in payroll_rows:
        te = time_entries[row["name"]]
        by_campaign = {c["campaign_name"]: c for c in row["campaign_breakdown"]}

        values = [
            te.agent_name, te.break_hours, te.training_hours, te.lunch_hours, te.manual_dial_hours,
            te.talk_hours, te.wait_hours, te.wrap_hours, row["manual_hours"], te.total_hours,
            row["hourly_rate"], row["hours_pay"],
        ]
        values += [by_campaign.get(c.name, {}).get("appointments", 0) for c in campaigns]
        values += [by_campaign.get(c.name, {}).get("valid_sits", 0) for c in campaigns]
        values += [by_campaign.get(c.name, {}).get("sales", 0) for c in campaigns]
        values += [row["total_appointments"], row["commission_total"], row["spiffs"], row["gross_pay"]]
        ws.append(values)

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).font = body_font
        for money_col in [11, 12, len(headers) - 2, len(headers) - 1, len(headers)]:
            ws.cell(row=row_idx, column=money_col).number_format = money_fmt
        row_idx += 1

    total_col = len(headers)
    ws.cell(row=row_idx + 1, column=total_col - 1, value="TOTAL PAYOUT").font = Font(name="Arial", bold=True, size=10)
    total_cell = ws.cell(row=row_idx + 1, column=total_col,
                          value=f"=SUM({get_column_letter(total_col)}2:{get_column_letter(total_col)}{row_idx - 1})")
    total_cell.font = Font(name="Arial", bold=True, size=10)
    total_cell.number_format = money_fmt

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(header) * 0.9)
    ws.freeze_panes = "B2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"WarpLine_Payroll_{period.label.replace(' ', '_').replace(',', '')}.xlsx"
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      as_attachment=True, download_name=filename)


@app.route("/period/<int:period_id>/paystubs.zip")
@login_required
def download_paystubs(period_id):
    period = PayPeriod.query.get_or_404(period_id)
    payroll_rows, _ = get_agent_results(period)

    if not payroll_rows:
        flash("No payroll data to generate paystubs from yet.", "error")
        return redirect(url_for("results_page", period_id=period.id))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in payroll_rows:
            docx_buffer = generate_paystub_docx(row["name"], period.label, row)
            safe_name = row["name"].replace(" ", "_").replace("/", "-")
            zf.writestr(f"{safe_name}.docx", docx_buffer.read())

    zip_buffer.seek(0)
    filename = f"WarpLine_Paystubs_{period.label.replace(' ', '_').replace(',', '')}.zip"
    return send_file(zip_buffer, mimetype="application/zip", as_attachment=True, download_name=filename)


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
